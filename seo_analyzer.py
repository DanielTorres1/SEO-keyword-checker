#!/usr/bin/env python3
"""
Script de análisis SEO que verifica la presencia de keywords en elementos clave de las páginas web.
Lee configuración desde site.config y genera un reporte en Excel con formato condicional.
Utiliza Selenium para manejar páginas dinámicas.
"""

import json
import requests
from bs4 import BeautifulSoup
from urllib.parse import urlparse, parse_qs
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import sys
import os
import time
import argparse
from datetime import datetime
from dotenv import load_dotenv

# Selenium imports
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager


def load_api_key():
    """Carga la API key desde el archivo .env si existe."""
    env_path = os.path.join(os.path.dirname(__file__), '.env')
    if os.path.exists(env_path):
        load_dotenv(env_path)
        return os.getenv('API_KEY')
    return None


def load_config(config_file='site.config'):
    """Carga el archivo de configuración JSON."""
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Error: No se encontró el archivo {config_file}")
        sys.exit(1)
    except json.JSONDecodeError:
        print(f"Error: El archivo {config_file} no es un JSON válido")
        sys.exit(1)


def is_url_friendly(url):
    """Verifica si la URL es amigable (sin parámetros GET)."""
    parsed = urlparse(url)
    # Retorna True si no hay parámetros de consulta
    return len(parse_qs(parsed.query)) == 0


def fetch_page_selenium(url):
    """Obtiene el contenido HTML de una URL usando Selenium para manejar contenido dinámico."""
    driver = None
    try:
        # Configurar Selenium
        options = webdriver.ChromeOptions()
        options.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option("useAutomationExtension", False)
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--headless")  # Ejecutar en modo headless
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        
        # Use system chromium binary
        options.binary_location = "/usr/bin/chromium"
        
        # Try to use system chromedriver first, fallback to ChromeDriverManager
        try:
            driver = webdriver.Chrome(options=options)
        except Exception:
            # Fallback to ChromeDriverManager
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options)
        driver.set_page_load_timeout(30)
        
        print(f"        Cargando página con Selenium...", end=' ', flush=True)
        driver.get(url)
        time.sleep(5)  # Esperar a que cargue el contenido dinámico
        
        # Obtener HTML renderizado
        html = driver.execute_script("return document.documentElement.outerHTML;")
        print("✓")
        
        return html
        
    except Exception as e:
        print(f"✗ Error: {e}")
        return None
    finally:
        if driver:
            try:
                driver.quit()
            except:
                pass


def check_keyword_in_text(text, keyword):
    """Verifica si la keyword está presente en el texto (case-insensitive)."""
    if not text:
        return False
    return keyword.lower() in text.lower()


def get_pagespeed_insights(url, strategy='mobile', api_key=None, retry=False):
    """Obtiene métricas de PageSpeed Insights para una URL.
    
    Args:
        url: URL a analizar
        strategy: 'mobile' o 'desktop'
        api_key: API key de Google (opcional, pero recomendado para evitar límites)
        retry: Si es True, indica que es un reintento
    
    Returns:
        dict con 'accessibility' y 'performance' scores (0-100), o None en caso de error
    """
    try:
        # URL de la API de PageSpeed Insights
        api_url = 'https://www.googleapis.com/pagespeedonline/v5/runPagespeed'
        
        params = {
            'url': url,
            'strategy': strategy,
            'category': ['ACCESSIBILITY', 'PERFORMANCE']
        }
        
        if api_key:
            params['key'] = api_key
        
        retry_text = " (reintento)" if retry else ""
        print(f"        Obteniendo PageSpeed ({strategy}){retry_text}...", end=' ', flush=True)
        response = requests.get(api_url, params=params, timeout=30)
        response.raise_for_status()
        
        data = response.json()
        
        # Extraer scores de las categorías
        lighthouse = data.get('lighthouseResult', {})
        categories = lighthouse.get('categories', {})
        
        accessibility_score = categories.get('accessibility', {}).get('score')
        performance_score = categories.get('performance', {}).get('score')
        
        # Convertir de 0-1 a 0-100
        accessibility = int(accessibility_score * 100) if accessibility_score is not None else None
        performance = int(performance_score * 100) if performance_score is not None else None
        
        print(f"✓ (Acc: {accessibility}, Perf: {performance})")
        
        # Pausa más larga para evitar rate limiting (sin API key)
        time.sleep(3 if not api_key else 1)
        
        return {
            'accessibility': accessibility,
            'performance': performance
        }
        
    except requests.RequestException as e:
        print(f"✗ Error: {e}")
        return {'accessibility': None, 'performance': None}
    except Exception as e:
        print(f"✗ Error inesperado: {e}")
        return {'accessibility': None, 'performance': None}


def analyze_page(url, keyword, include_pagespeed=True, api_key=None):
    """Analiza una página web y verifica la presencia de la keyword en elementos SEO.
    
    Args:
        url: URL a analizar
        keyword: Palabra clave a buscar
        include_pagespeed: Si True, incluye análisis de PageSpeed Insights
        api_key: API key de Google para PageSpeed (opcional)
    """
    html = fetch_page_selenium(url)
    
    if not html:
        return {
            'URL': url,
            'Keyword': keyword,
            'Título': 'ERROR',
            'Keyword en Título': False,
            'Meta Descripción': 'ERROR',
            'Keyword en Meta Descripción': False,
            'H1': 'ERROR',
            'Keyword en H1': False,
            'Alt Text Imágenes': 'ERROR',
            'Keyword en Alt Text': False,
            'Keyword en Alt Ratio': '0 de 0',
            'Alt Images With Keyword': 0,
            'Alt Total Images': 0,
            'URL Amigable': is_url_friendly(url),
            'Accesibilidad(Mobile)': None,
            'PageSpeed(Mobile)': None,
            'Accesibilidad(Web)': None,
            'PageSpeed(Web)': None
        }
    
    soup = BeautifulSoup(html, 'html.parser')
    
    # 1. Título
    title_tag = soup.find('title')
    title = title_tag.get_text(strip=True) if title_tag else ''
    keyword_in_title = check_keyword_in_text(title, keyword)
    
    # 2. Meta descripción
    meta_desc_tag = soup.find('meta', attrs={'name': 'description'})
    meta_desc = meta_desc_tag.get('content', '') if meta_desc_tag else ''
    keyword_in_meta = check_keyword_in_text(meta_desc, keyword)
    
    # 3. H1 - Use separator to preserve spaces between nested elements
    h1_tags = soup.find_all('h1')
    h1_text = ' | '.join([h1.get_text(separator=' ', strip=True) for h1 in h1_tags])
    keyword_in_h1 = check_keyword_in_text(h1_text, keyword)
    
    # 4. Alt text de imágenes
    img_tags = soup.find_all('img')
    alt_texts = [img.get('alt', '') for img in img_tags if img.get('alt')]
    alt_text_combined = ' | '.join(alt_texts)
    
    # Contar imágenes con keyword en alt text
    total_images_with_alt = len(alt_texts)
    images_with_keyword = sum(1 for alt in alt_texts if check_keyword_in_text(alt, keyword))
    keyword_in_alt_ratio = f"{images_with_keyword} de {total_images_with_alt}" if total_images_with_alt > 0 else "0 de 0"
    keyword_in_alt = any(check_keyword_in_text(alt, keyword) for alt in alt_texts)
    
    # 5. URL amigable
    url_friendly = is_url_friendly(url)
    
    # 6. PageSpeed Insights (Mobile y Desktop)
    mobile_metrics = {'accessibility': None, 'performance': None}
    desktop_metrics = {'accessibility': None, 'performance': None}
    
    if include_pagespeed:
        mobile_metrics = get_pagespeed_insights(url, 'mobile', api_key)
        desktop_metrics = get_pagespeed_insights(url, 'desktop', api_key)
        
        # Reintentar si los resultados de desktop son None
        if desktop_metrics['accessibility'] is None or desktop_metrics['performance'] is None:
            print("        ⚠ Resultados de desktop incompletos, reintentando...")
            desktop_metrics = get_pagespeed_insights(url, 'desktop', api_key, retry=True)
    
    return {
        'URL': url,
        'Keyword': keyword,
        'Título': title,
        'Keyword en Título': keyword_in_title,
        'Meta Descripción': meta_desc,
        'Keyword en Meta Descripción': keyword_in_meta,
        'H1': h1_text if h1_text else 'No H1 encontrado',
        'Keyword en H1': keyword_in_h1,
        'Alt Text Imágenes': alt_text_combined if alt_text_combined else 'Sin alt text',
        'Keyword en Alt Text': keyword_in_alt,
        'Keyword en Alt Ratio': keyword_in_alt_ratio,
        'Alt Images With Keyword': images_with_keyword,
        'Alt Total Images': total_images_with_alt,
        'URL Amigable': url_friendly,
        'Accesibilidad(Mobile)': mobile_metrics['accessibility'],
        'PageSpeed(Mobile)': mobile_metrics['performance'],
        'Accesibilidad(Web)': desktop_metrics['accessibility'],
        'PageSpeed(Web)': desktop_metrics['performance']
    }


def create_excel_report(results, output_file='seo_report.xlsx'):
    """Crea un reporte en Excel con formato condicional."""
    # Crear libro de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Análisis SEO"
    
    # Definir colores
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Encabezados
    headers = [
        'URL', 'Keyword', 
        'Título', '✓ Keyword en Título',
        'Meta Descripción', '✓ Keyword en Meta',
        'H1', '✓ Keyword en H1',
        'Alt Text Imágenes', 'Keyword en Alt Text',
        '✓ URL Amigable',
        'Accesibilidad(Mobile)', 'PageSpeed(Mobile)',
        'Accesibilidad(Web)', 'PageSpeed(Web)'
    ]
    
    ws.append(headers)
    
    # Hacer los encabezados en negrita
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Agregar datos
    for result in results:
        row = [
            result['URL'],
            result['Keyword'],
            result['Título'],
            'SÍ' if result['Keyword en Título'] else 'NO',
            result['Meta Descripción'],
            'SÍ' if result['Keyword en Meta Descripción'] else 'NO',
            result['H1'],
            'SÍ' if result['Keyword en H1'] else 'NO',
            result['Alt Text Imágenes'],
            result['Keyword en Alt Ratio'],  # Mostrar ratio en lugar de SÍ/NO
            'SÍ' if result['URL Amigable'] else 'NO',
            result['Accesibilidad(Mobile)'] if result['Accesibilidad(Mobile)'] is not None else 'N/A',
            result['PageSpeed(Mobile)'] if result['PageSpeed(Mobile)'] is not None else 'N/A',
            result['Accesibilidad(Web)'] if result['Accesibilidad(Web)'] is not None else 'N/A',
            result['PageSpeed(Web)'] if result['PageSpeed(Web)'] is not None else 'N/A'
        ]
        ws.append(row)
    
    # Aplicar formato condicional (colorear celdas de verificación)
    check_columns = [4, 6, 8, 11]  # Columnas D, F, H, K (sin J porque ahora es ratio)
    
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in check_columns:
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value == 'SÍ':
                cell.fill = green_fill
            elif cell.value == 'NO':
                cell.fill = red_fill
    
    # Aplicar formato condicional a la columna J (Keyword en Alt - ratio)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=10)  # Columna J
        cell_value = str(cell.value)
        
        # Extraer números del formato "X de Y"
        if ' de ' in cell_value:
            try:
                parts = cell_value.split(' de ')
                with_keyword = int(parts[0])
                total = int(parts[1])
                
                # Verde si más de la mitad tienen keyword, rojo si no
                if total > 0:
                    if with_keyword > total / 2:
                        cell.fill = green_fill
                    else:
                        cell.fill = red_fill
            except (ValueError, IndexError):
                pass  # Si no se puede parsear, no colorear
    
    # Aplicar formato condicional a las columnas de PageSpeed (L, M, N, O)
    pagespeed_columns = [12, 13, 14, 15]  # Columnas L, M, N, O
    
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in pagespeed_columns:
            cell = ws.cell(row=row_idx, column=col_idx)
            # Solo colorear si el valor es numérico
            if isinstance(cell.value, (int, float)):
                score = cell.value
                if 0 <= score <= 49:
                    cell.fill = red_fill
                elif 50 <= score <= 89:
                    cell.fill = yellow_fill
                elif 90 <= score <= 100:
                    cell.fill = green_fill
    
    # Ajustar ancho de columnas
    column_widths = {
        'A': 50,  # URL
        'B': 30,  # Keyword
        'C': 50,  # Título
        'D': 18,  # Check Título
        'E': 50,  # Meta Descripción
        'F': 18,  # Check Meta
        'G': 50,  # H1
        'H': 18,  # Check H1
        'I': 50,  # Alt Text
        'J': 18,  # Check Alt
        'K': 15,  # URL Amigable
        'L': 20,  # Accesibilidad Mobile
        'M': 18,  # PageSpeed Mobile
        'N': 20,  # Accesibilidad Web
        'O': 18   # PageSpeed Web
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Guardar archivo
    wb.save(output_file)
    print(f"\n✓ Reporte generado: {output_file}")


def main():
    """Función principal."""
    parser = argparse.ArgumentParser(
        description='Analizador SEO - Verifica keywords y métricas de rendimiento',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos de uso:
  %(prog)s                                     # Análisis completo con PageSpeed
  %(prog)s --no-pagespeed                      # Solo análisis SEO sin PageSpeed
  %(prog)s --api-key YOUR_KEY                  # Con API key de Google
  %(prog)s --config-file custom.config         # Usar archivo de config personalizado
  %(prog)s --config-file sites.json --api-key YOUR_KEY  # Personalizado con API key
        """
    )
    
    parser.add_argument(
        '--config-file',
        type=str,
        default='site.config',
        help='Archivo de configuración JSON (default: site.config)'
    )
    
    parser.add_argument(
        '--no-pagespeed',
        action='store_true',
        help='Desactivar análisis de PageSpeed Insights (más rápido)'
    )
    
    parser.add_argument(
        '--api-key',
        type=str,
        default=None,
        help='API key de Google para PageSpeed Insights (evita límites de tasa)'
    )
    
    args = parser.parse_args()
    
    # Extraer variables de los argumentos
    include_pagespeed = not args.no_pagespeed
    # Load API key from .env if not provided via command line
    api_key = args.api_key or load_api_key()
    
    print("=" * 60)
    print("  ANALIZADOR SEO - Verificación de Keywords")
    print("=" * 60)
    
    print(f"\nCargando configuración desde: {args.config_file}")
    config = load_config(args.config_file)
    
    print(f"URLs a analizar: {len(config)}")
    if include_pagespeed:
        print("PageSpeed Insights: ACTIVADO")
        if api_key:
            print("  └─ Usando API key (sin límites de tasa)")
        else:
            print("  └─ Sin API key (puede ser lento, límites de tasa aplicados)")
            print("  └─ Consigue una API key gratis: https://developers.google.com/speed/docs/insights/v5/get-started")
    else:
        print("PageSpeed Insights: DESACTIVADO")
    print()
    
    # Analizar cada página
    results = []
    for idx, site in enumerate(config, 1):
        url = site.get('URL')
        
        # Manejar keywords como array o string
        keywords_raw = site.get('keywords', site.get('Keyword'))
        if isinstance(keywords_raw, list):
            keywords_list = keywords_raw
        elif keywords_raw:
            keywords_list = [keywords_raw]
        else:
            keywords_list = ['']
        
        print(f"[{idx}/{len(config)}] Analizando: {url}")
        keywords_formatted = ', '.join([f"'{k}'" for k in keywords_list])
        print(f"    Keywords: {keywords_formatted}")
        
        # Analizar la página una vez y reutilizar para todos los keywords
        # Primero obtenemos el HTML y PageSpeed solo una vez
        first_result = analyze_page(url, keywords_list[0], include_pagespeed, api_key)
        
        # Para cada keyword, crear un resultado
        for keyword in keywords_list:
            if keyword == keywords_list[0]:
                # Ya tenemos el resultado del primer keyword
                result = first_result
            else:
                # Para los demás keywords, reutilizar PageSpeed pero recalcular checks de keyword
                result = analyze_page(url, keyword, include_pagespeed=False, api_key=api_key)
                # Copiar métricas de PageSpeed del primer resultado
                result['Accesibilidad(Mobile)'] = first_result['Accesibilidad(Mobile)']
                result['PageSpeed(Mobile)'] = first_result['PageSpeed(Mobile)']
                result['Accesibilidad(Web)'] = first_result['Accesibilidad(Web)']
                result['PageSpeed(Web)'] = first_result['PageSpeed(Web)']
            
            results.append(result)
        
        # Mostrar resumen agregado
        total_checks = 0
        total_passed = 0
        for keyword in keywords_list:
            # Encontrar el resultado para este keyword
            for r in results:
                if r['URL'] == url and r['Keyword'] == keyword:
                    checks = [
                        r['Keyword en Título'],
                        r['Keyword en Meta Descripción'],
                        r['Keyword en H1'],
                        r['Keyword en Alt Text'],
                        r['URL Amigable']
                    ]
                    passed = sum(checks)
                    total_checks += 5
                    total_passed += passed
                    print(f"      '{keyword}': {passed}/5 verificaciones pasadas")
                    break
        
        print(f"    ✓ Total: {total_passed}/{total_checks} verificaciones pasadas\n")
    
    # Generar reporte Excel
    # Generar reporte Excel con fecha en el nombre
    current_date = datetime.now().strftime('%Y-%m-%d')
    output_file = f'seo_report_{current_date}.xlsx'
    create_excel_report(results, output_file)
    
    print("=" * 60)
    print(f"Análisis completado. Total URLs: {len(results)}")
    print("=" * 60)


if __name__ == '__main__':
    main()
